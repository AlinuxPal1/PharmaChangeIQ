from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.dimensions import ColumnDimension


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "data" / "output" / "PharmaChangeIQ_Assessed.xlsx"


TAB_COLORS = {
    "Instructions": "CBD5E1",
    "Change_Intake": "93C5FD",
    "Risk_Matrix": "FDBA74",
    "Regulatory_Impact": "BFDBFE",
    "QA_Documentation": "C4B5FD",
    "Decision_Summary": "FDE68A",
    "Portfolio_Assessment": "A7F3D0",
    "Dashboard": "86EFAC",
}


def apply_common_layout(ws):
    ws.sheet_view.showGridLines = False

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color="111827")
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if ws.max_row >= 1:
        ws.row_dimensions[1].height = 30
        ws["A1"].font = Font(name="Aptos", size=18, bold=True, color="111827")

    if ws.max_row >= 2:
        ws.row_dimensions[2].height = 42
        ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="4B5563")
        ws["A2"].alignment = Alignment(vertical="top", wrap_text=True)


def style_header_row(ws, row=4):
    fill = PatternFill("solid", fgColor="E8EEF7")
    border = Border(bottom=Side(style="thin", color="B8C2CC"))

    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = Font(name="Aptos", size=10, bold=True, color="111827")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    ws.row_dimensions[row].height = 30


def normalize_body_rows(ws, start_row=5, end_row=None, height=24):
    if end_row is None:
        end_row = ws.max_row

    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = height


def polish_instructions(ws):
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 120

    for row in range(6, 11):
        ws.row_dimensions[row].height = 34

    ws.freeze_panes = "A6"


def polish_change_intake(ws):
    widths = {
        "A": 14,
        "B": 30,
        "C": 22,
        "D": 12,
        "E": 30,
        "F": 58,
        "G": 20,
        "H": 18,
        "I": 24,
        "J": 20,
        "K": 24,
        "L": 34,
        "M": 16,
        "N": 20,
        "O": 18,
        "P": 42,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    style_header_row(ws, 4)
    normalize_body_rows(ws, 5, 20, 30)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:P54"


def polish_risk_matrix(ws):
    widths = {
        "A": 16,
        "B": 32,
        "C": 12,
        "D": 12,
        "E": 18,
        "F": 90,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    style_header_row(ws, 4)
    normalize_body_rows(ws, 5, 20, 34)
    ws.freeze_panes = "A5"


def polish_regulatory_impact(ws):
    widths = {
        "A": 16,
        "B": 72,
        "C": 18,
        "D": 18,
        "E": 82,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    style_header_row(ws, 4)
    normalize_body_rows(ws, 5, 20, 36)
    ws.freeze_panes = "A5"


def polish_qa_documentation(ws):
    widths = {
        "A": 16,
        "B": 42,
        "C": 14,
        "D": 32,
        "E": 88,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    style_header_row(ws, 4)
    normalize_body_rows(ws, 5, 25, 34)
    ws.freeze_panes = "A5"


def polish_decision_summary(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 115

    for row in range(4, 15):
        ws.row_dimensions[row].height = 34
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F3F4F6")
        ws.cell(row=row, column=1).font = Font(name="Aptos", size=10, bold=True, color="111827")

    ws.row_dimensions[13].height = 78
    ws.row_dimensions[14].height = 88
    ws.freeze_panes = "A4"


def polish_portfolio_assessment(ws):
    widths = {
        "A": 14,
        "B": 32,
        "C": 22,
        "D": 12,
        "E": 32,
        "F": 16,
        "G": 22,
        "H": 20,
        "I": 30,
        "J": 30,
        "K": 100,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    style_header_row(ws, 4)
    normalize_body_rows(ws, 5, 20, 42)
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:K104"


def polish_dashboard(ws):
    widths = {
        "A": 34,
        "B": 24,
        "C": 4,
        "D": 36,
        "E": 18,
        "F": 4,
        "G": 28,
        "H": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(4, 20):
        ws.row_dimensions[row].height = 26

    for cell in ws[4]:
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor="E8EEF7")
            cell.font = Font(name="Aptos", size=10, bold=True, color="111827")

    ws.freeze_panes = "A4"


def main():
    if not OUTPUT_PATH.exists():
        raise FileNotFoundError(f"Output workbook not found: {OUTPUT_PATH}")

    wb = load_workbook(OUTPUT_PATH)

    for ws in wb.worksheets:
        apply_common_layout(ws)

        if ws.title in TAB_COLORS:
            ws.sheet_properties.tabColor = TAB_COLORS[ws.title]

    if "Instructions" in wb.sheetnames:
        polish_instructions(wb["Instructions"])

    if "Change_Intake" in wb.sheetnames:
        polish_change_intake(wb["Change_Intake"])

    if "Risk_Matrix" in wb.sheetnames:
        polish_risk_matrix(wb["Risk_Matrix"])

    if "Regulatory_Impact" in wb.sheetnames:
        polish_regulatory_impact(wb["Regulatory_Impact"])

    if "QA_Documentation" in wb.sheetnames:
        polish_qa_documentation(wb["QA_Documentation"])

    if "Decision_Summary" in wb.sheetnames:
        polish_decision_summary(wb["Decision_Summary"])

    if "Portfolio_Assessment" in wb.sheetnames:
        polish_portfolio_assessment(wb["Portfolio_Assessment"])

    if "Dashboard" in wb.sheetnames:
        polish_dashboard(wb["Dashboard"])

    if "Dashboard" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Dashboard")

    wb.save(OUTPUT_PATH)
    print(f"Workbook polished successfully: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()

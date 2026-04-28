from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


RISK_AREA_ORDER = [
    "Regulatory Impact",
    "QA Impact",
    "Safety Impact",
    "Labelling Impact",
    "Supply Chain Impact",
    "Business Continuity Impact",
]


AREA_WEIGHTS = {
    "Regulatory Impact": "30%",
    "QA Impact": "25%",
    "Safety Impact": "20%",
    "Labelling Impact": "15%",
    "Supply Chain Impact": "5%",
    "Business Continuity Impact": "5%",
}


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = None


def _apply_output_style(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    border = Border(bottom=Side(style="thin", color="E5E7EB"))
    font = Font(name="Aptos", size=10, color="111827")

    for row in range(min_row, max_row + 1):
        ws.row_dimensions[row].height = 28
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _classification_fill(classification: str) -> PatternFill:
    if classification == "Critical Change":
        return PatternFill("solid", fgColor="FCA5A5")
    if classification == "Major Change":
        return PatternFill("solid", fgColor="FDBA74")
    if classification == "Moderate Change":
        return PatternFill("solid", fgColor="FDE68A")
    if classification == "Minor Change":
        return PatternFill("solid", fgColor="BBF7D0")
    return PatternFill("solid", fgColor="E5E7EB")


def write_assessment_results(
    template_path: str | Path,
    output_path: str | Path,
    intake_df,
    assessments,
) -> Path:
    """
    Write risk assessment results into a copy of the Excel workbook.

    The original template is preserved. Results are saved to output_path.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    if not assessments:
        raise ValueError("No assessments provided.")

    wb = load_workbook(template_path)

    # MVP: one assessed case at a time in formatted output sheets.
    assessment = assessments[0]
    intake_row = intake_df.iloc[0]

    # -------------------------
    # Risk_Matrix
    # -------------------------
    ws = wb["Risk_Matrix"]
    _clear_range(ws, 5, 30, 1, 6)

    start_row = 5
    for idx, area in enumerate(RISK_AREA_ORDER, start=start_row):
        score = assessment.area_scores[area]
        weight = AREA_WEIGHTS[area]
        weighted_score = round(score * float(weight.strip("%")) / 100, 2)
        rationale = assessment.rationales[area]

        ws.cell(row=idx, column=1).value = assessment.change_id
        ws.cell(row=idx, column=2).value = area
        ws.cell(row=idx, column=3).value = score
        ws.cell(row=idx, column=4).value = weight
        ws.cell(row=idx, column=5).value = weighted_score
        ws.cell(row=idx, column=6).value = rationale

    summary_row = start_row + len(RISK_AREA_ORDER) + 2

    summary_items = [
        ("Overall Weighted Score", assessment.weighted_score),
        ("Overall Classification", assessment.classification),
        ("Escalation Required", assessment.escalation_required),
        ("Primary Owner", assessment.primary_owner),
        ("Secondary Owner", assessment.secondary_owner),
    ]

    for offset, (label, value) in enumerate(summary_items):
        row = summary_row + offset
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=1).font = Font(name="Aptos", size=10, bold=True)
        ws.cell(row=row, column=2).font = Font(name="Aptos", size=10, bold=True)

    ws.cell(row=summary_row + 1, column=2).fill = _classification_fill(assessment.classification)

    _apply_output_style(ws, 5, summary_row + len(summary_items), 1, 6)

    # -------------------------
    # Regulatory_Impact
    # -------------------------
    ws = wb["Regulatory_Impact"]
    _clear_range(ws, 5, 30, 1, 5)

    regulatory_rows = [
        (
            "Does the change affect product claims?",
            "Yes" if intake_row["Change Type"] == "Claim Update" else "No",
            "High" if intake_row["Change Type"] == "Claim Update" else "Low",
            "Claim-related changes require substantiation and regulatory review.",
        ),
        (
            "Does the change affect label or artwork?",
            intake_row["Label Impacted"],
            "High" if intake_row["Label Impacted"] == "Yes" else "Low",
            "Label impact requires controlled artwork/labelling review.",
        ),
        (
            "Does the change involve an ingredient?",
            intake_row["Ingredient Involved"],
            "Moderate" if intake_row["Ingredient Involved"] == "Yes" else "Low",
            "Ingredient involvement may affect compliance or safety evaluation.",
        ),
        (
            "Does the change affect existing regulatory approval or notification?",
            intake_row["Existing Regulatory Approval Impacted"],
            "High" if intake_row["Existing Regulatory Approval Impacted"] in {"Yes", "Probably"} else "Moderate",
            "RA confirmation is required when approval impact is confirmed, likely, or uncertain.",
        ),
        (
            "Is market-specific review required?",
            "Yes",
            "Moderate",
            f"Selected market: {intake_row['Market']}. Market-specific requirements should be considered.",
        ),
    ]

    for row_idx, item in enumerate(regulatory_rows, start=5):
        question, answer, impact, comment = item
        ws.cell(row=row_idx, column=1).value = assessment.change_id
        ws.cell(row=row_idx, column=2).value = question
        ws.cell(row=row_idx, column=3).value = answer
        ws.cell(row=row_idx, column=4).value = impact
        ws.cell(row=row_idx, column=5).value = comment

    _apply_output_style(ws, 5, 5 + len(regulatory_rows), 1, 5)

    # -------------------------
    # QA_Documentation
    # -------------------------
    ws = wb["QA_Documentation"]
    _clear_range(ws, 5, 35, 1, 5)

    required_documents = [
        (
            "Change Control Form",
            "Yes",
            "QA",
            "A controlled change record is required for all assessed changes.",
        ),
        (
            "Regulatory Impact Assessment",
            "Yes" if assessment.area_scores["Regulatory Impact"] >= 2 else "No",
            "Regulatory Affairs",
            "Required when regulatory impact is moderate or higher.",
        ),
        (
            "Label/Artwork Review",
            "Yes" if assessment.area_scores["Labelling Impact"] >= 2 else "No",
            "Regulatory Affairs / QA",
            "Required when label, artwork, claims, warnings, or instructions are impacted.",
        ),
        (
            "Safety Assessment",
            "Yes" if assessment.area_scores["Safety Impact"] >= 2 else "Possibly",
            "Safety / Toxicology",
            "Required or considered when ingredient, exposure, or safety data issues are present.",
        ),
        (
            "Supplier Qualification",
            "Yes" if intake_row["Supplier Impacted"] == "Yes" else "No",
            "QA / Supply Chain",
            "Required when supplier impact is identified.",
        ),
        (
            "Stability Assessment",
            "Yes" if intake_row["Change Type"] == "Stability Related Change" else "No",
            "R&D / QA",
            "Required for changes potentially affecting stability, shelf life, or storage conditions.",
        ),
        (
            "Final Approval Memo",
            "Yes",
            "QA / Regulatory Affairs",
            "Required to document final decision, ownership, and implementation conditions.",
        ),
    ]

    for row_idx, item in enumerate(required_documents, start=5):
        document, required, owner, rationale = item
        ws.cell(row=row_idx, column=1).value = assessment.change_id
        ws.cell(row=row_idx, column=2).value = document
        ws.cell(row=row_idx, column=3).value = required
        ws.cell(row=row_idx, column=4).value = owner
        ws.cell(row=row_idx, column=5).value = rationale

    _apply_output_style(ws, 5, 5 + len(required_documents), 1, 5)


    # -------------------------
    # Portfolio_Assessment
    # -------------------------
    if "Portfolio_Assessment" not in wb.sheetnames:
        wb.create_sheet("Portfolio_Assessment")

    ws = wb["Portfolio_Assessment"]

    # Ensure headers exist
    portfolio_headers = [
        "Change ID",
        "Product Name",
        "Product Category",
        "Market",
        "Change Type",
        "Weighted Score",
        "Classification",
        "Escalation Required",
        "Primary Owner",
        "Secondary Owner",
        "Key Driver",
    ]

    for col_idx, header in enumerate(portfolio_headers, start=1):
        ws.cell(row=4, column=col_idx).value = header

    _clear_range(ws, 5, 200, 1, len(portfolio_headers))

    for row_offset, current_assessment in enumerate(assessments, start=5):
        current_intake = intake_df[intake_df["Change ID"] == current_assessment.change_id].iloc[0]

        key_driver_area = max(
            current_assessment.area_scores,
            key=current_assessment.area_scores.get,
        )

        key_driver = (
            f"{key_driver_area}: score {current_assessment.area_scores[key_driver_area]}. "
            f"{current_assessment.rationales[key_driver_area]}"
        )

        values = [
            current_assessment.change_id,
            current_intake["Product Name"],
            current_intake["Product Category"],
            current_intake["Market"],
            current_intake["Change Type"],
            current_assessment.weighted_score,
            current_assessment.classification,
            current_assessment.escalation_required,
            current_assessment.primary_owner,
            current_assessment.secondary_owner,
            key_driver,
        ]

        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_offset, column=col_idx).value = value

        ws.cell(row=row_offset, column=7).fill = _classification_fill(current_assessment.classification)

    _apply_output_style(ws, 5, 5 + len(assessments), 1, len(portfolio_headers))

    # -------------------------
    # Decision_Summary
    # -------------------------
    ws = wb["Decision_Summary"]

    summary_values = {
        "Change ID": assessment.change_id,
        "Product Name": intake_row["Product Name"],
        "Product Category": intake_row["Product Category"],
        "Market": intake_row["Market"],
        "Change Type": intake_row["Change Type"],
        "Overall Risk Classification": assessment.classification,
        "Escalation Required": assessment.escalation_required,
        "Primary Owner": assessment.primary_owner,
        "Secondary Owner": assessment.secondary_owner,
        "Main Rationale": (
            f"The proposed change was classified as {assessment.classification}. "
            f"Main drivers include regulatory score {assessment.area_scores['Regulatory Impact']} "
            f"and labelling score {assessment.area_scores['Labelling Impact']}. "
            f"{assessment.rationales['Regulatory Impact']}"
        ),
        "Recommended Actions": (
            "1. Perform functional RA/QA review before implementation. "
            "2. Confirm required documentation package. "
            "3. Document final decision and approval owner. "
            "4. Escalate if market authorization, safety, or labelling obligations are confirmed."
        ),
    }

    for row in range(4, 15):
        label = ws.cell(row=row, column=1).value
        if label in summary_values:
            ws.cell(row=row, column=2).value = summary_values[label]

    ws["B9"].fill = _classification_fill(assessment.classification)
    _apply_output_style(ws, 4, 14, 1, 2)


    # -------------------------
    # Dashboard
    # -------------------------
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard")

    ws = wb["Dashboard"]

    # Clear dashboard area
    _clear_range(ws, 4, 50, 1, 8)

    total_changes = len(assessments)
    major_changes = sum(1 for a in assessments if a.classification == "Major Change")
    critical_changes = sum(1 for a in assessments if a.classification == "Critical Change")
    escalation_required = sum(1 for a in assessments if a.escalation_required == "Yes")
    average_score = round(sum(a.weighted_score for a in assessments) / total_changes, 2)

    change_types = intake_df["Change Type"].value_counts()
    most_frequent_change_type = change_types.index[0] if not change_types.empty else "N/A"

    primary_owners = [a.primary_owner for a in assessments]
    most_involved_owner = max(set(primary_owners), key=primary_owners.count) if primary_owners else "N/A"

    dashboard_rows = [
        ["Metric", "Value"],
        ["Total changes assessed", total_changes],
        ["Major changes", major_changes],
        ["Critical changes", critical_changes],
        ["Escalation required", escalation_required],
        ["Average weighted score", average_score],
        ["Most frequent change type", most_frequent_change_type],
        ["Most involved primary owner", most_involved_owner],
    ]

    for row_idx, row_values in enumerate(dashboard_rows, start=4):
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    ws.cell(row=4, column=1).font = Font(name="Aptos", size=10, bold=True)
    ws.cell(row=4, column=2).font = Font(name="Aptos", size=10, bold=True)

    ws.cell(row=12, column=1).value = "Classification distribution"
    ws.cell(row=12, column=1).font = Font(name="Aptos", size=11, bold=True)

    classification_counts = {}
    for assessment_item in assessments:
        classification_counts[assessment_item.classification] = classification_counts.get(assessment_item.classification, 0) + 1

    ws.cell(row=13, column=1).value = "Classification"
    ws.cell(row=13, column=2).value = "Count"
    ws.cell(row=13, column=1).font = Font(name="Aptos", size=10, bold=True)
    ws.cell(row=13, column=2).font = Font(name="Aptos", size=10, bold=True)

    for row_offset, (classification, count) in enumerate(classification_counts.items(), start=14):
        ws.cell(row=row_offset, column=1).value = classification
        ws.cell(row=row_offset, column=2).value = count
        ws.cell(row=row_offset, column=1).fill = _classification_fill(classification)

    ws.cell(row=12, column=4).value = "Primary owner distribution"
    ws.cell(row=12, column=4).font = Font(name="Aptos", size=11, bold=True)

    owner_counts = {}
    for assessment_item in assessments:
        owner_counts[assessment_item.primary_owner] = owner_counts.get(assessment_item.primary_owner, 0) + 1

    ws.cell(row=13, column=4).value = "Primary Owner"
    ws.cell(row=13, column=5).value = "Count"
    ws.cell(row=13, column=4).font = Font(name="Aptos", size=10, bold=True)
    ws.cell(row=13, column=5).font = Font(name="Aptos", size=10, bold=True)

    for row_offset, (owner, count) in enumerate(owner_counts.items(), start=14):
        ws.cell(row=row_offset, column=4).value = owner
        ws.cell(row=row_offset, column=5).value = count

    _apply_output_style(ws, 4, 30, 1, 8)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path

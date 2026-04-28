from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "templates" / "PharmaChangeIQ_Template.xlsx"


SHEETS = [
    "Instructions",
    "Change_Intake",
    "Risk_Matrix",
    "Regulatory_Impact",
    "QA_Documentation",
    "Decision_Summary",
    "Portfolio_Assessment",
    "Dashboard",
]


VALIDATION_OPTIONS = {
    "Product Category": [
        "Pharma/OTC",
        "Food Supplement",
        "Cosmetic",
        "Medical Device",
        "Consumer Health",
    ],
    "Market": ["EU", "US", "UK", "CH", "Global"],
    "Change Type": [
        "Ingredient Change",
        "Claim Update",
        "Label Artwork Change",
        "Packaging Supplier Change",
        "Manufacturing Site Change",
        "Process Change",
        "Raw Material Supplier Change",
        "Specification Change",
        "Stability Related Change",
        "Market Expansion",
    ],
    "Yes No Partial": ["Yes", "No", "Partial", "Unknown", "Probably"],
    "Urgency": ["Low", "Medium", "High", "Critical"],
}


def style_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos", size=18, bold=True, color="1F2937")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="4B5563")
        ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[2].height = 34


def style_header_row(ws, row, start_col, end_col):
    fill = PatternFill("solid", fgColor="E8EEF7")
    font = Font(name="Aptos", size=10, bold=True, color="1F2937")
    border = Border(bottom=Side(style="thin", color="B8C2CC"))

    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_body_range(ws, min_row, max_row, min_col, max_col):
    border = Border(bottom=Side(style="thin", color="E5E7EB"))
    font = Font(name="Aptos", size=10, color="111827")

    for row in range(min_row, max_row + 1):
        ws.row_dimensions[row].height = 22
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_dropdown(ws, cell_range, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def setup_instructions(ws):
    style_title(
        ws,
        "PharmaChangeIQ",
        "Excel-based regulatory and QA change impact assessment template. Fill the Change_Intake sheet, then run the Python backend to generate risk scoring, documentation mapping, and decision outputs.",
    )

    rows = [
        ["Step", "Action"],
        ["1", "Fill one row in Change_Intake for each proposed product, process, supplier, packaging, labelling, or claim-related change."],
        ["2", "Use predefined dropdown values where available to keep inputs controlled and auditable."],
        ["3", "Run the Python backend using: python run_assessment.py"],
        ["4", "Review Risk_Matrix, Regulatory_Impact, QA_Documentation, and Decision_Summary."],
        ["5", "Export or share the generated decision summary as part of the portfolio case study."],
    ]

    for r_idx, row in enumerate(rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    style_header_row(ws, 5, 1, 2)
    style_body_range(ws, 6, 10, 1, 2)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A6"


def setup_change_intake(ws):
    style_title(
        ws,
        "Change Intake",
        "Operational input form for proposed changes. Each row represents one change scenario to be assessed.",
    )

    headers = [
        "Change ID",
        "Product Name",
        "Product Category",
        "Market",
        "Change Type",
        "Change Description",
        "Ingredient Involved",
        "Label Impacted",
        "Manufacturing Impacted",
        "Supplier Impacted",
        "Safety Data Available",
        "Existing Regulatory Approval Impacted",
        "Urgency",
        "Submitted By",
        "Submission Date",
        "Notes",
    ]

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header

    style_header_row(ws, 4, 1, len(headers))
    style_body_range(ws, 5, 54, 1, len(headers))

    widths = {
        "A": 14,
        "B": 24,
        "C": 22,
        "D": 14,
        "E": 28,
        "F": 48,
        "G": 20,
        "H": 18,
        "I": 24,
        "J": 20,
        "K": 24,
        "L": 34,
        "M": 16,
        "N": 20,
        "O": 18,
        "P": 38,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    add_dropdown(ws, "C5:C54", VALIDATION_OPTIONS["Product Category"])
    add_dropdown(ws, "D5:D54", VALIDATION_OPTIONS["Market"])
    add_dropdown(ws, "E5:E54", VALIDATION_OPTIONS["Change Type"])
    add_dropdown(ws, "G5:L54", VALIDATION_OPTIONS["Yes No Partial"])
    add_dropdown(ws, "M5:M54", VALIDATION_OPTIONS["Urgency"])

    sample_rows = [
        [
            "CHG-001",
            "Example Immune Support Supplement",
            "Food Supplement",
            "EU",
            "Claim Update",
            "Add a new immune-support claim to the product label.",
            "No",
            "Yes",
            "No",
            "No",
            "Partial",
            "Probably",
            "Medium",
            "Portfolio user",
            "2026-04-28",
            "Example case: claim update for a food supplement.",
        ],
        [
            "CHG-002",
            "Example OTC Oral Solution",
            "Pharma/OTC",
            "EU",
            "Packaging Supplier Change",
            "Change the supplier of the primary bottle used for an OTC oral solution.",
            "No",
            "No",
            "No",
            "Yes",
            "Yes",
            "Probably",
            "High",
            "Portfolio user",
            "2026-04-28",
            "Example case: primary packaging supplier change for regulated product.",
        ],
        [
            "CHG-003",
            "Example Facial Serum",
            "Cosmetic",
            "EU",
            "Ingredient Change",
            "Replace fragrance component due to allergen and consumer tolerability concerns.",
            "Yes",
            "Yes",
            "No",
            "No",
            "Partial",
            "Unknown",
            "Medium",
            "Portfolio user",
            "2026-04-28",
            "Example case: cosmetic fragrance replacement.",
        ],
    ]

    for row_offset, sample_row in enumerate(sample_rows, start=5):
        for idx, value in enumerate(sample_row, start=1):
            ws.cell(row=row_offset, column=idx).value = value

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:P54"


def setup_risk_matrix(ws):
    style_title(
        ws,
        "Risk Matrix",
        "Risk scoring output generated by the Python backend. Values shown here are initially placeholders.",
    )

    headers = ["Change ID", "Risk Area", "Score", "Weight", "Weighted Score", "Rationale"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header

    areas = [
        ["CHG-001", "Regulatory Impact", "", "30%", "", ""],
        ["CHG-001", "QA Impact", "", "25%", "", ""],
        ["CHG-001", "Safety Impact", "", "20%", "", ""],
        ["CHG-001", "Labelling Impact", "", "15%", "", ""],
        ["CHG-001", "Supply Chain Impact", "", "5%", "", ""],
        ["CHG-001", "Business Continuity Impact", "", "5%", "", ""],
    ]

    for r_idx, row in enumerate(areas, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    ws["A13"] = "Overall Classification"
    ws["B13"] = "Pending backend assessment"
    ws["A14"] = "Escalation Required"
    ws["B14"] = "Pending backend assessment"
    ws["A15"] = "Primary Owner"
    ws["B15"] = "Pending backend assessment"
    ws["A16"] = "Secondary Owner"
    ws["B16"] = "Pending backend assessment"

    style_header_row(ws, 4, 1, 6)
    style_body_range(ws, 5, 16, 1, 6)

    widths = {"A": 16, "B": 30, "C": 12, "D": 12, "E": 18, "F": 70}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


def setup_regulatory_impact(ws):
    style_title(
        ws,
        "Regulatory Impact",
        "Structured regulatory review questions and backend-generated impact interpretation.",
    )

    headers = ["Change ID", "Question", "Answer", "Impact", "Comment"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header

    questions = [
        ["CHG-001", "Does the change affect product claims?", "", "", ""],
        ["CHG-001", "Does the change affect label or artwork?", "", "", ""],
        ["CHG-001", "Does the change involve a regulated or restricted ingredient?", "", "", ""],
        ["CHG-001", "Does the change affect existing authorization, notification, or compliance status?", "", "", ""],
        ["CHG-001", "Does the change introduce a new market-specific requirement?", "", "", ""],
    ]

    for r_idx, row in enumerate(questions, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    style_header_row(ws, 4, 1, 5)
    style_body_range(ws, 5, 20, 1, 5)

    widths = {"A": 16, "B": 70, "C": 16, "D": 18, "E": 70}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


def setup_qa_documentation(ws):
    style_title(
        ws,
        "QA Documentation",
        "Document mapping output showing which controlled documents or reviews are required for each assessed change.",
    )

    headers = ["Change ID", "Required Document", "Required", "Function Owner", "Rationale"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header

    documents = [
        ["CHG-001", "Change Control Form", "", "QA", ""],
        ["CHG-001", "Regulatory Impact Assessment", "", "Regulatory Affairs", ""],
        ["CHG-001", "Label/Artwork Review", "", "Regulatory Affairs / QA", ""],
        ["CHG-001", "Safety Assessment", "", "Safety / Toxicology", ""],
        ["CHG-001", "Supplier Qualification", "", "QA / Supply Chain", ""],
        ["CHG-001", "Stability Assessment", "", "R&D / QA", ""],
        ["CHG-001", "Final Approval Memo", "", "QA / Regulatory Affairs", ""],
    ]

    for r_idx, row in enumerate(documents, start=5):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    style_header_row(ws, 4, 1, 5)
    style_body_range(ws, 5, 25, 1, 5)

    widths = {"A": 16, "B": 34, "C": 14, "D": 30, "E": 75}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"


def setup_decision_summary(ws):
    style_title(
        ws,
        "Decision Summary",
        "Management-ready change assessment summary generated by the Python backend.",
    )

    labels = [
        "Change ID",
        "Product Name",
        "Product Category",
        "Market",
        "Change Type",
        "Overall Risk Classification",
        "Escalation Required",
        "Primary Owner",
        "Secondary Owner",
        "Main Rationale",
        "Recommended Actions",
    ]

    for idx, label in enumerate(labels, start=4):
        ws.cell(row=idx, column=1).value = label
        ws.cell(row=idx, column=2).value = "Pending backend assessment"

    style_body_range(ws, 4, 14, 1, 2)

    for row in range(4, 15):
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F3F4F6")
        ws.cell(row=row, column=1).font = Font(name="Aptos", size=10, bold=True, color="1F2937")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A4"



def setup_portfolio_assessment(ws):
    style_title(
        ws,
        "Portfolio Assessment",
        "Multi-case summary generated by the Python backend. Each row represents one assessed change case.",
    )

    headers = [
        "Change ID",
        "Product Name",
        "Product Category",
        "Market",
        "Change Type",
        "Weighted Score",
        "Classification",
        "Escalation Required",
        "Primary Owner",
        "Secondary Owner",
        "Key Driver",
    ]

    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx).value = header

    style_header_row(ws, 4, 1, len(headers))
    style_body_range(ws, 5, 104, 1, len(headers))

    widths = {
        "A": 14,
        "B": 30,
        "C": 22,
        "D": 12,
        "E": 30,
        "F": 16,
        "G": 22,
        "H": 20,
        "I": 28,
        "J": 28,
        "K": 80,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:K104"

def setup_dashboard(ws):
    style_title(
        ws,
        "Dashboard",
        "Portfolio-level visual summary. This sheet will be populated after multiple change cases are assessed.",
    )

    kpis = [
        ["Metric", "Value"],
        ["Total changes assessed", "Pending"],
        ["High/Critical changes", "Pending"],
        ["Most frequent change type", "Pending"],
        ["Most involved function", "Pending"],
    ]

    for r_idx, row in enumerate(kpis, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    style_header_row(ws, 4, 1, 2)
    style_body_range(ws, 5, 8, 1, 2)

    ws["D4"] = "Dashboard area reserved for charts generated from assessed cases."
    ws["D4"].font = Font(name="Aptos", size=11, italic=True, color="4B5563")
    ws["D4"].alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["D"].width = 70


def build_workbook():
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    for sheet_name in SHEETS:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

    setup_instructions(wb["Instructions"])
    setup_change_intake(wb["Change_Intake"])
    setup_risk_matrix(wb["Risk_Matrix"])
    setup_regulatory_impact(wb["Regulatory_Impact"])
    setup_qa_documentation(wb["QA_Documentation"])
    setup_decision_summary(wb["Decision_Summary"])
    setup_portfolio_assessment(wb["Portfolio_Assessment"])
    setup_dashboard(wb["Dashboard"])

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template created: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_workbook()
